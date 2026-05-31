from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any, Dict, List


def _configure_stdio_utf8() -> None:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            pass


def apply_patches(xlsx_path: Path, patch_doc: Dict[str, Any], out_path: Path) -> int:
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError("missing openpyxl; install: pip install openpyxl") from e

    patches = patch_doc.get("patches")
    if not isinstance(patches, list):
        raise ValueError("patch document must contain patches array")

    wb = openpyxl.load_workbook(xlsx_path)
    applied = 0

    for i, p in enumerate(patches):
        if not isinstance(p, dict):
            raise ValueError(f"patches[{i}] must be an object")
        sheet_name = str(p.get("sheet", wb.active.title))
        row = int(p["row"])
        column = str(p["column"]).strip()
        new_value = p.get("new")
        if row < 2:
            raise ValueError(f"patches[{i}].row must be >= 2 (row 1 is header)")
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"unknown sheet: {sheet_name!r}")
        ws = wb[sheet_name]
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        col_idx = None
        for idx, h in enumerate(headers, start=1):
            if h == column:
                col_idx = idx
                break
        if col_idx is None:
            raise ValueError(f"column {column!r} not found in sheet {sheet_name!r}; headers={headers}")
        ws.cell(row=row, column=col_idx, value=new_value)
        applied += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.resolve() == xlsx_path.resolve():
        raise ValueError(
            "refusing to overwrite input xlsx; use a different output path "
            f"(e.g. {xlsx_path.with_stem(xlsx_path.stem + '_patched')})"
        )
    wb.save(out_path)
    return applied


def _default_output_path(xlsx_in: Path) -> Path:
    """Never overwrite source workbook."""
    stem = xlsx_in.stem
    if stem.endswith("_patched"):
        return xlsx_in.with_name(f"{stem}_1{xlsx_in.suffix}")
    return xlsx_in.with_name(f"{stem}_patched{xlsx_in.suffix}")


def main(argv: List[str]) -> int:
    _configure_stdio_utf8()
    if len(argv) not in (3, 4):
        print(
            "usage: python apply_excel_patch.py <patches.json> <input.xlsx> [output.xlsx]\n"
            "  If output.xlsx is omitted, writes <input>_patched.xlsx (never overwrites input)."
        )
        return 2
    patch_path = Path(argv[1]).expanduser().resolve()
    xlsx_in = Path(argv[2]).expanduser().resolve()
    xlsx_out = Path(argv[3]).expanduser().resolve() if len(argv) == 4 else _default_output_path(xlsx_in)
    doc = json.loads(patch_path.read_text(encoding="utf-8"))
    n = apply_patches(xlsx_in, doc, xlsx_out)
    print(f"applied {n} patch(es) -> {xlsx_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
